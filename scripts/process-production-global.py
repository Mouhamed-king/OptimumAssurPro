#!/usr/bin/env python3
"""Nettoie Production_Global et recalcule les primes nettes via les baremes auto."""

from __future__ import annotations

import argparse
import math
import shutil
import sys
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from cotation_loader import CotationData, load_cotation, _normalize_category, _parse_duration_months

COLUMNS_TO_REMOVE = {
    "Nom Point",
    "Compagnie",
    "CODEQUAL",
    "DATE EMISSION",
    "BONUS RC",
    "REDUCTION COMMERCIALE",
    "GENRE",
    "USAGE",
    "MEC",
    "PLACES",
    "CYLINDRE",
    "VALEUR NEUF",
    "VALEUR VENALE",
    "GARANTIES",
    "FORMIPT",
    "FORMAVR",
    "NUMERO ATTESTATION",
    "NUMERO CEDEAO",
    "NOM CONDUCTEUR",
    "DATE NAISSANCE",
    "ACCESSOIRE",
    "Taxes",
    "FGA",
    "Prime TTC",
}

REQUIRED_COLUMNS = {
    "CATEGORIE",
    "DUREE",
    "PUISSANCE",
    "CHARGE UTILE",
    "ENERGIE",
    "Prime Nette",
}


def _header_map(worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=col).value
        if value is not None and str(value).strip():
            headers[str(value).strip()] = col
    return headers


def _get_cell(row_values: dict[str, object], headers: dict[str, int], name: str):
    col = headers.get(name)
    if col is None:
        return None
    return row_values.get(col)


def _row_to_dict(worksheet, row_idx: int, headers: dict[str, int]) -> dict[int, object]:
    return {col: worksheet.cell(row=row_idx, column=col).value for col in headers.values()}


def _set_cell(worksheet, row_idx: int, headers: dict[str, int], name: str, value) -> None:
    col = headers.get(name)
    if col is not None:
        worksheet.cell(row=row_idx, column=col, value=value)


def calculate_prime_nette(
    cotation: CotationData,
    categorie: str,
    duree: str,
    puissance,
    charge_utile,
    energie,
) -> tuple[float | None, str | None]:
    cat = _normalize_category(categorie)
    months = _parse_duration_months(duree)
    if months is None:
        return None, f"Durée non reconnue: {duree!r}"

    try:
        cv = int(float(puissance or 0))
    except (TypeError, ValueError):
        cv = 0

    rc_annual: float | None = None

    if cotation.is_tpc_tariff(categorie, str(charge_utile or "")):
        if cv <= 0:
            return None, "Puissance invalide pour TPC"
        prime = cotation.prime_for_tpc(cv, duree)
        if prime is None:
            return None, f"Aucun tarif TPC pour puissance {cv} et duree {duree!r}"
        return prime, None

    if cat == "1":
        if cv <= 0:
            return None, "Puissance invalide pour CAT 1"
        rc_annual = cotation.rc_for_cat1(cv)
        if rc_annual is None:
            return None, f"Aucune tranche CV pour puissance {cv} (CAT 1)"

    elif cat == "2":
        if cv <= 0:
            return None, "Puissance invalide pour CAT 2"
        rc_annual = cotation.rc_for_cat2(cv, str(energie or ""), str(charge_utile or ""))
        if rc_annual is None:
            return None, (
                f"Aucune tranche pour CAT 2 (puissance={cv}, energie={energie!r}, "
                f"charge utile={charge_utile!r})"
            )

    elif cat == "5":
        # Regle metier: toute CAT 5 est traitee au tarif cyclomoteur,
        # quelle que soit la cylindree renseignee dans le fichier source.
        rc_annual = cotation.rc_for_cat5()
        if rc_annual is None:
            return None, "Tarif cyclomoteur introuvable pour CAT 5"

    else:
        return None, f"Catégorie non gérée: {categorie!r}"

    prime = cotation.prime_nette_bonus(rc_annual, duree)
    if prime is None:
        return None, f"Durée {duree!r} absente du tableau bonus 20%"

    return prime, None


def remove_columns(worksheet, headers: dict[str, int]) -> dict[str, int]:
    cols_to_delete = sorted(
        (col for name, col in headers.items() if name in COLUMNS_TO_REMOVE),
        reverse=True,
    )
    for col in cols_to_delete:
        worksheet.delete_cols(col)

    return _header_map(worksheet)


def process_workbook(
    production_path: Path,
    cotation_path: Path,
    dry_run: bool = False,
) -> dict[str, object]:
    if not production_path.exists():
        raise FileNotFoundError(f"Fichier introuvable: {production_path}")
    if not cotation_path.exists():
        raise FileNotFoundError(f"Cotation introuvable: {cotation_path}")

    cotation = load_cotation(cotation_path)
    workbook = openpyxl.load_workbook(production_path)
    worksheet = workbook.active

    headers = _header_map(worksheet)
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        workbook.close()
        raise ValueError(f"Colonnes manquantes: {', '.join(sorted(missing))}")

    stats = {
        "processed": 0,
        "updated": 0,
        "unchanged": 0,
        "warnings": [],
    }

    for row_idx in range(2, worksheet.max_row + 1):
        categorie = worksheet.cell(row=row_idx, column=headers["CATEGORIE"]).value
        if categorie is None or str(categorie).strip() == "":
            continue

        stats["processed"] += 1
        row_values = _row_to_dict(worksheet, row_idx, headers)

        prime, warning = calculate_prime_nette(
            cotation=cotation,
            categorie=str(categorie),
            duree=str(_get_cell(row_values, headers, "DUREE") or ""),
            puissance=_get_cell(row_values, headers, "PUISSANCE"),
            charge_utile=_get_cell(row_values, headers, "CHARGE UTILE"),
            energie=_get_cell(row_values, headers, "ENERGIE"),
        )

        client = _get_cell(row_values, headers, "NOM-PRENOM Assuré") or f"ligne {row_idx}"
        if warning:
            stats["warnings"].append(f"{client}: {warning}")
            continue

        old_prime = _get_cell(row_values, headers, "Prime Nette")
        try:
            old_value = int(math.floor(float(old_prime))) if old_prime is not None else None
        except (TypeError, ValueError):
            old_value = None

        new_value = int(prime)
        _set_cell(worksheet, row_idx, headers, "Prime Nette", new_value)

        if old_value == new_value:
            stats["unchanged"] += 1
        else:
            stats["updated"] += 1

    headers = remove_columns(worksheet, headers)

    if not dry_run:
        workbook.save(production_path)

    workbook.close()
    return stats


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Nettoie Production_Global et recalcule les primes nettes (bonus 20%)."
    )
    parser.add_argument(
        "production_file",
        nargs="?",
        default=None,
        help="Chemin vers le fichier Production_Global .xlsx",
    )
    parser.add_argument(
        "--cotation",
        default=str(PROJECT_ROOT / "data" / "baremes-assurance.json"),
        help="Chemin vers le fichier global des baremes (.json) ou Cotation Auto.xls",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Calcule sans sauvegarder le fichier",
    )
    args = parser.parse_args()

    if args.production_file:
        production_path = Path(args.production_file)
    else:
        matches = sorted(PROJECT_ROOT.glob("Production_Global*.xlsx"))
        if not matches:
            print("Erreur: aucun fichier Production_Global*.xlsx trouvé.", file=sys.stderr)
            return 1
        if len(matches) > 1:
            print("Plusieurs fichiers trouvés, précisez le chemin:", file=sys.stderr)
            for match in matches:
                print(f"  - {match.name}", file=sys.stderr)
            return 1
        production_path = matches[0]

    cotation_path = Path(args.cotation)

    print(f"Fichier production : {production_path}")
    print(f"Cotation           : {cotation_path}")

    stats = process_workbook(production_path, cotation_path, dry_run=args.dry_run)

    print(f"\nLignes traitées    : {stats['processed']}")
    print(f"Primes mises à jour: {stats['updated']}")
    print(f"Primes inchangées  : {stats['unchanged']}")

    warnings = stats["warnings"]
    if warnings:
        print(f"\nAvertissements ({len(warnings)}):")
        for warning in warnings:
            print(f"  - {warning}")

    if args.dry_run:
        print("\nMode dry-run: aucune modification enregistrée.")
    else:
        print(f"\nFichier enregistré : {production_path}")

    return 1 if warnings else 0


if __name__ == "__main__":
    raise SystemExit(main())
